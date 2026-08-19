#!/usr/bin/env python3
"""Build the Final Adjudication queue for the site.

Reads the researcher Excel (Adjudication Queue + Task Details sheets) and emits
the minimum the page needs:

  data/final_adjudication/queue.json            queue entries
  data/final_adjudication/images/FA_XXXX.jpg    GT-only render (always)
  data/final_adjudication/images_model/FA_XXXX.jpg  existing ghost vis (if any)

The Excel is never copied to the site and is never modified. Reviewer-facing
JSON carries only what the adjudicator must read: the queue id, the issue, the
previous decisions and their notes, and the image paths. Condition labels,
reviewer identities, primary/R3 task ids and original file names stay out of it
-- they are recoverable offline from the Excel via queue_id.

Model-assisted views reuse the EXISTING ghost visualisations. No inference runs
here and no visualisation is regenerated.

Requires: openpyxl, Pillow.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import shutil
import subprocess
import sys
from collections import Counter
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from gt_render import load_coco, load_font, render_gt_only  # noqa: E402

SITE_ROOT = Path(__file__).resolve().parents[1]
# Location of the research repository. Override with EPINU_RESEARCH_ROOT;
# defaults to a sibling checkout so no absolute personal path is hard-coded.
RESEARCH_ROOT = Path(
    os.environ.get("EPINU_RESEARCH_ROOT", SITE_ROOT.parent / "epinu-rfdetr-training")
)

DEFAULT_XLSX = SITE_ROOT / "data" / "Claim2_Human_Ghost_Audit_Final_Confirmation_List.xlsx"
DEFAULT_AUDIT = RESEARCH_ROOT / "data/claim2/ghost_audit/audit"
DEFAULT_GHOST_VIS = RESEARCH_ROOT / "data/human_ghost_audit/model_guided_primary"

QUEUE_SHEET = "Adjudication Queue"
DETAIL_SHEET = "Task Details"
EXPECTED_QUEUE = 26

ISSUE_EXPLANATION = {
    "YES vs NO disagreement":
        "The primary reviewer thought a defect was present, but Reviewer 3 did not.",
    "Defect-type mismatch":
        "Both reviews found a defect, but they disagreed about the defect type.",
}

results: list[tuple[str, str, bool, str]] = []


def check(section: str, name: str, ok: bool, detail: str = "") -> bool:
    results.append((section, name, bool(ok), detail))
    return bool(ok)


def die(message: str) -> None:
    print(f"\nERROR: {message}", file=sys.stderr)
    sys.exit(1)


def sheet_rows(ws) -> list[dict]:
    header = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2):
        values = [c.value for c in row]
        if all(v is None for v in values):
            continue
        rows.append(dict(zip(header, values)))
    return rows


def text(value) -> str:
    """Excel cell -> trimmed string ('' for blanks)."""
    if value is None:
        return ""
    return str(value).strip()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--audit-dir", type=Path, default=DEFAULT_AUDIT)
    parser.add_argument("--ghost-vis-dir", type=Path, default=DEFAULT_GHOST_VIS)
    parser.add_argument("--force", action="store_true")
    args = parser.parse_args()

    if not args.xlsx.is_file():
        die(
            f"Final Confirmation workbook not found: {args.xlsx}\n"
            "Not inventing a queue -- the selection is fixed. Point --xlsx at it."
        )
    images_dir = args.audit_dir / "images"
    coco_path = args.audit_dir / "_annotations.coco.json"
    for path in (images_dir, coco_path):
        if not path.exists():
            die(f"input not found: {path}")

    out_dir = SITE_ROOT / "data" / "final_adjudication"
    gt_dir, model_dir = out_dir / "images", out_dir / "images_model"
    if gt_dir.exists() and any(gt_dir.iterdir()) and not args.force:
        die(
            f"{gt_dir} already exists and is not empty.\n"
            "Refusing to overwrite in case an adjudication is in progress. "
            "Re-run with --force to rebuild."
        )
    if out_dir.exists():
        shutil.rmtree(out_dir)
    gt_dir.mkdir(parents=True)
    model_dir.mkdir(parents=True)

    # The workbook is opened read-only and never written back.
    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    for sheet in (QUEUE_SHEET, DETAIL_SHEET):
        if sheet not in wb.sheetnames:
            die(f"sheet '{sheet}' not found in {args.xlsx.name}")

    queue_rows = [r for r in sheet_rows(wb[QUEUE_SHEET]) if r.get("queue_id") is not None]
    detail_rows = [r for r in sheet_rows(wb[DETAIL_SHEET]) if text(r.get("task_id"))]

    # Primary reviews, grouped by source image. Task Details holds one row per
    # primary task, so an image reviewed under both conditions has two.
    primaries: dict[str, list[dict]] = {}
    for row in detail_rows:
        primaries.setdefault(text(row["original_file_name"]), []).append(row)

    image_by_name, anns_by_image = load_coco(coco_path)
    font = load_font(13)

    entries = []
    mapping = []
    unresolved: list[tuple[str, str]] = []
    model_views = 0

    for index, row in enumerate(queue_rows, start=1):
        queue_id = row["queue_id"]
        original = text(row["original_file_name"])
        image_id = f"FA_{index:04d}"

        src = images_dir / original
        meta = image_by_name.get(original)
        if not src.is_file() or meta is None:
            unresolved.append((str(queue_id), "source image or COCO entry missing"))
            continue
        render_gt_only(src, anns_by_image[meta["id"]], gt_dir / f"{image_id}.jpg", font)

        # Reuse the existing ghost visualisation when one exists for this source.
        vis_src = args.ghost_vis_dir / (Path(original).stem + "_eval.jpg")
        model_image = None
        if vis_src.is_file():
            shutil.copy2(vis_src, model_dir / f"{image_id}.jpg")
            model_image = f"images_model/{image_id}.jpg"
            model_views += 1

        issue = text(row["issue"])
        primary_reviews = [
            {
                "decision": text(p["primary_defect_found"]),
                "defect_types": text(p["primary_defect_types"]),
                "target_classes": text(p["primary_target_classes"]),
                "number_of_defects": text(p["primary_number_of_defects"]),
                "notes": text(p["primary_notes"]),
            }
            for p in primaries.get(original, [])
        ]
        if not primary_reviews:
            unresolved.append((str(queue_id), "no primary review found in Task Details"))
            continue

        entries.append({
            "queue_id": queue_id,
            "image_id": image_id,
            "issue": issue,
            "issue_explanation": ISSUE_EXPLANATION.get(issue, ""),
            "gt_image": f"images/{image_id}.jpg",
            "model_image": model_image,
            "primary_reviews": primary_reviews,
            "reviewer3": {
                "decision": text(row["r3_defect_found"]),
                "defect_types": text(row["r3_defect_types"]),
                "target_classes": text(row["r3_target_classes"]),
                "number_of_defects": text(row["r3_number_of_defects"]),
                "notes": text(row["r3_notes"]),
            },
        })
        mapping.append({
            "queue_id": queue_id,
            "image_id": image_id,
            "original_file_name": original,
        })

    if unresolved:
        print("\nFAIL: unresolved queue entries")
        for queue_id, reason in unresolved:
            print(f"  queue_id={queue_id}: {reason}")

    (out_dir / "queue.json").write_text(
        json.dumps({"queue_count": len(entries), "entries": entries}, indent=2) + "\n",
        encoding="utf-8",
    )

    # Researcher-only: queue_id -> original file name, for merging results back.
    mapping_path = SITE_ROOT / "data" / "final_adjudication_mapping_researcher.csv"
    with mapping_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=["queue_id", "image_id", "original_file_name"])
        writer.writeheader()
        writer.writerows(mapping)

    # ---------------------------------------------------------------- checks
    queue_ids = [e["queue_id"] for e in entries]
    originals = [m["original_file_name"] for m in mapping]

    check("queue", f"queue entries == {EXPECTED_QUEUE}", len(entries) == EXPECTED_QUEUE,
          f"got {len(entries)}")
    check("queue", "queue_id values match the workbook exactly",
          queue_ids == [r["queue_id"] for r in queue_rows])
    check("queue", "no duplicate queue_id", len(set(queue_ids)) == len(queue_ids))
    check("queue", "no duplicate source image (one final decision per image)",
          len(set(originals)) == len(originals),
          f"{len(set(originals))} unique / {len(originals)}")
    check("queue", "unresolved == 0", not unresolved,
          f"{len(unresolved)} unresolved" if unresolved else "")
    # A source image reviewed under both conditions can appear in the queue via
    # one primary task and in QA Misses via the other -- that is expected. What
    # must not happen is a QA-miss TASK being adjudicated here.
    qa_rows = [r for r in sheet_rows(wb["QA Misses"]) if text(r.get("task_id"))]
    qa_tasks = {text(r["task_id"]) for r in qa_rows}
    queue_tasks = {t.strip() for r in queue_rows
                   for t in str(r["primary_task_ids"]).split(";") if t.strip()}
    check("queue", "no QA-miss task appears in the adjudication queue",
          not (qa_tasks & queue_tasks),
          f"shared tasks: {sorted(qa_tasks & queue_tasks)}")
    shared_images = set(originals) & {text(r["original_file_name"]) for r in qa_rows}
    check("queue", "QA Misses remain a separate list",
          len(qa_rows) == 2,
          f"{len(qa_rows)} QA-miss rows; {len(shared_images)} share a source image "
          f"with the queue via a different primary task")

    check("previous_reviews", "every entry has at least one primary review",
          all(e["primary_reviews"] for e in entries))
    multi = [e for e in entries if len(e["primary_reviews"]) > 1]
    check("previous_reviews", "images with two primary tasks keep both reviews",
          len(multi) == sum(1 for r in queue_rows
                            if len(str(r["primary_task_ids"]).split(";")) > 1),
          f"{len(multi)} entries with 2 reviews")
    check("previous_reviews", "total primary reviews == Task Details rows",
          sum(len(e["primary_reviews"]) for e in entries) == len(detail_rows),
          f"{sum(len(e['primary_reviews']) for e in entries)} vs {len(detail_rows)}")
    check("previous_reviews", "every entry has a Reviewer 3 decision",
          all(e["reviewer3"]["decision"] for e in entries))
    check("previous_reviews", "every issue has an explanation",
          all(e["issue_explanation"] for e in entries),
          str({e["issue"] for e in entries if not e["issue_explanation"]}))

    gt_files = sorted(p.name for p in gt_dir.iterdir() if p.is_file())
    model_files = sorted(p.name for p in model_dir.iterdir() if p.is_file())
    check("images", f"GT-only images == {EXPECTED_QUEUE}", len(gt_files) == EXPECTED_QUEUE,
          f"got {len(gt_files)}")
    check("images", "GT image names match image ids",
          gt_files == sorted(f"{e['image_id']}.jpg" for e in entries))
    check("images", "model-assisted views only where a visualisation exists",
          len(model_files) == model_views and
          model_files == sorted(f"{e['image_id']}.jpg" for e in entries if e["model_image"]),
          f"{model_views} of {EXPECTED_QUEUE}")
    check("images", "model views are byte-identical copies of the existing ghost vis",
          all(
              (model_dir / f"{m['image_id']}.jpg").read_bytes()
              == (args.ghost_vis_dir / (Path(m["original_file_name"]).stem + "_eval.jpg")).read_bytes()
              for m in mapping
              if (model_dir / f"{m['image_id']}.jpg").is_file()
          ))
    check("images", "source images untouched",
          all((images_dir / o).is_file() for o in originals))

    payload = (out_dir / "queue.json").read_text(encoding="utf-8")

    # Everything except the free-text notes the adjudicator is meant to read.
    structural = json.dumps([
        {k: v for k, v in e.items() if k not in ("primary_reviews", "reviewer3")}
        | {
            "primary_reviews": [
                {k: v for k, v in pr.items() if k != "notes"}
                for pr in e["primary_reviews"]
            ],
            "reviewer3": {k: v for k, v in e["reviewer3"].items() if k != "notes"},
        }
        for e in entries
    ])

    check("privacy", "no original file name in queue.json",
          not any(o in payload for o in originals))
    for column in ["conditions", "primary_reviewers", "primary_task_ids", "r3_task_id"]:
        values = {text(r[column]) for r in queue_rows if text(r[column])}
        check("privacy", f"no {column} value in queue.json structure",
              not any(v in structural for v in values))
    for term in ["random", "ghost", "condition", "reviewer1", "reviewer2",
                 "hga_", "r3_", "original_file"]:
        check("privacy", f"queue.json structure contains no '{term}'",
              term not in structural.lower())
    # Notes are shown verbatim by design; assert only that they carry no identifiers.
    notes_text = " ".join(
        [pr["notes"] for e in entries for pr in e["primary_reviews"]]
        + [e["reviewer3"]["notes"] for e in entries]
    )
    check("privacy", "reviewer notes contain no task id or file name",
          not any(o in notes_text for o in originals)
          and "HGA_" not in notes_text and "R3_" not in notes_text)

    # git check-ignore -q takes a single pathname, so ask about each in turn.
    def git_ignored(path: Path) -> bool:
        return subprocess.run(
            ["git", "check-ignore", "-q", str(path.relative_to(SITE_ROOT))],
            cwd=SITE_ROOT, capture_output=True,
        ).returncode == 0

    check("privacy", "researcher mapping is git-ignored", git_ignored(mapping_path))
    check("privacy", "workbook is git-ignored", git_ignored(args.xlsx))

    section = None
    for sec, name, ok, detail in results:
        if sec != section:
            print(f"\n[{sec}]")
            section = sec
        line = f"  {'PASS' if ok else 'FAIL'}  {name}"
        if detail:
            line += f"  -- {detail}"
        print(line)

    failed = [r for r in results if not r[2]]
    print("\n" + "=" * 60)
    print(f"queue entries      : {len(entries)}")
    print(f"GT-only images     : {len(gt_files)}")
    print(f"model-assisted     : {model_views}")
    print(f"issues             : {dict(Counter(e['issue'] for e in entries))}")
    print(f"researcher mapping : {mapping_path}")
    if failed:
        print(f"RESULT: FAIL ({len(failed)} of {len(results)})")
        return 1
    print(f"RESULT: PASS ({len(results)} checks)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
